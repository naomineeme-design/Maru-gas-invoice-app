import streamlit as st
from PIL import Image
import openpyxl
import easyocr
import re

st.set_page_config(page_title="請求書アプリ", layout="centered")

# ====================== パスワード ======================
PASSWORD = "komuro2026"   # ← ここは自由に変更してください

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

if not st.session_state.authenticated:
    st.title("🔐 丸磯請求書アプリ")
    pw = st.text_input("パスワードを入力してください", type="password")
    if st.button("ログイン", use_container_width=True):
        if pw == PASSWORD:
            st.session_state.authenticated = True
            st.rerun()
        else:
            st.error("パスワードが違います")
    st.stop()

# ====================== 本体 ======================
st.title("🛢️ ガソリン請求書作成アプリ")
st.caption("iPad・スマホ対応 | 手動入力対応版")

@st.cache_resource
def get_ocr():
    return easyocr.Reader(['ja', 'en'], gpu=False)

reader = get_ocr()

# 値を保持するためのセッション状態
if "tax10" not in st.session_state:
    st.session_state.tax10 = 0
if "keiyu" not in st.session_state:
    st.session_state.keiyu = 0

st.subheader("① 請求書の写真")
st.info("📸 iPadの場合：カメラ起動後、画面上部のカメラアイコンをタップすると背面カメラに切り替えられます")
camera = st.camera_input("カメラで撮影", key="cam")
upload = st.file_uploader("📁 写真を選択（こちらをおすすめ）", type=["jpg", "jpeg", "png"])

image = None
if camera:
    image = Image.open(camera)
elif upload:
    image = Image.open(upload)

if image:
    st.image(image, use_column_width=True)

    if st.button("🔍 OCRで自動読み取り", type="primary", use_container_width=True):
        with st.spinner("OCR処理中..."):
            text = " ".join(reader.readtext(image, detail=0))
            tax10_match = re.search(r'10%\s*対象合計[:：]?\s*([0-9,]+)', text) or re.search(r'税率\s*10%\s*対象[:：]?\s*([0-9,]+)', text)
            keiyu_match = re.search(r'軽油税合計[:：]?\s*([0-9,]+)', text)

            st.session_state.tax10 = int(tax10_match.group(1).replace(',', '')) if tax10_match else st.session_state.tax10
            st.session_state.keiyu = int(keiyu_match.group(1).replace(',', '')) if keiyu_match else st.session_state.keiyu

        st.success(f"OCR結果 → 10%対象: ¥{st.session_state.tax10:,}円　軽油税: ¥{st.session_state.keiyu:,}円")

    # ====================== 手動入力（常に表示） ======================
    st.subheader("② 金額を手動で入力・訂正")
    col1, col2 = st.columns(2)
    with col1:
        st.session_state.tax10 = st.number_input("10%対象税抜金額（必須）", value=st.session_state.tax10, step=1000)
    with col2:
        st.session_state.keiyu = st.number_input("軽油税金額（必須）", value=st.session_state.keiyu, step=100)

    st.subheader("③ 請求書情報")
    col_a, col_b = st.columns(2)
    with col_a:
        year = st.number_input("西暦", value=2026)
        month = st.number_input("月", value=3, min_value=1, max_value=12)
        day = st.number_input("日", value=15, min_value=1, max_value=31)
    with col_b:
        code = st.text_input("お客様コード", "07359-333200-057")
        work = st.text_input("工事名称", "大熊減容化作業所")

    if st.button("📄 Excelを作成する", type="primary", use_container_width=True):
        try:
            wb = openpyxl.load_workbook("2026  実験　丸磯　大熊減溶化作業所　.xlsx")
            ws = wb["入力シート・貴社控"]

            ws['F5'] = year
            ws['K5'] = month
            ws['N5'] = day
            ws['H8'] = work
            ws['B18'] = "別紙明細の通り"
            ws['O18'] = 10
            ws['Q18'] = "式"
            ws['AI18'] = 1
            ws['AM18'] = st.session_state.tax10
            ws['B20'] = "別紙明細の通り"
            ws['Q20'] = "式"
            ws['AI20'] = 1
            ws['AM20'] = st.session_state.keiyu

            filename = f"請求書_{year}{month:02d}{day:02d}.xlsx"
            wb.save(filename)

            with open(filename, "rb") as f:
                st.download_button("📥 Excelをダウンロード", f, filename, use_container_width=True)
            
            st.success("✅ Excelファイルを作成しました！ダウンロードしてください。")
        except FileNotFoundError:
            st.error("⚠️ Excelテンプレートが見つかりません。ファイル名が正確か確認してください。")
        except Exception as e:
            st.error(f"エラー: {e}")

st.caption("手動入力で確実に作成できます。OCRは補助的に使ってください。")
